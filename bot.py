import datetime
import io
import csv
import sqlite3
import os
import threading
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telebot import TeleBot, types
from apscheduler.schedulers.background import BackgroundScheduler

# --- Configuration & Initialization ---
API_TOKEN = os.environ.get('BOT_TOKEN')
ADMIN_CHAT_ID = int(os.environ.get('ADMIN_CHAT_ID', '0'))

if not API_TOKEN:
    raise ValueError("BOT_TOKEN environment variable is required!")

bot = TeleBot(API_TOKEN, parse_mode='Markdown')

scheduler = BackgroundScheduler()
scheduler.start()

DB_FILE = 'wash_and_scan.db'

# --- Health Check Server (for Render Web Service) ---
class HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b'OK')
    def log_message(self, format, *args):
        pass

def run_health_server():
    port = int(os.environ.get('PORT', 10000))
    server = HTTPServer(('0.0.0.0', port), HealthHandler)
    server.serve_forever()

# Start health check in background thread
threading.Thread(target=run_health_server, daemon=True).start()

# --- Database Initialization & Migration ---
def init_db():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_states (
                chat_id INTEGER PRIMARY KEY,
                state TEXT DEFAULT 'main_menu',
                pending_service TEXT,
                pending_price INTEGER,
                last_bot_msg_id INTEGER
            )
        ''')
        
        cursor.execute("PRAGMA table_info(user_states)")
        columns = [column[1] for column in cursor.fetchall()]
        if 'selected_cat_id' not in columns:
            cursor.execute("ALTER TABLE user_states ADD COLUMN selected_cat_id INTEGER")
            print("Added selected_cat_id to user_states")
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE,
                icon TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS services (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_id INTEGER,
                name TEXT,
                price INTEGER,
                is_active INTEGER DEFAULT 1,
                FOREIGN KEY(category_id) REFERENCES categories(id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id INTEGER,
                order_id_user INTEGER,
                category_id INTEGER,
                service TEXT,
                price INTEGER,
                time TEXT,
                date TEXT
            )
        ''')
        
        cursor.execute("PRAGMA table_info(orders)")
        order_columns = [column[1] for column in cursor.fetchall()]
        if 'category_id' not in order_columns:
            cursor.execute("ALTER TABLE orders ADD COLUMN category_id INTEGER")
            print("Added category_id to orders")
            
        conn.commit()

init_db()

# --- Helper Functions ---

def get_user_state(chat_id):
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT state, pending_service, pending_price, last_bot_msg_id, selected_cat_id FROM user_states WHERE chat_id = ?", (chat_id,))
        row = cursor.fetchone()
        
        if not row:
            cursor.execute("INSERT INTO user_states (chat_id) VALUES (?)", (chat_id,))
            conn.commit()
            return {"state": "main_menu", "pending_order": None, "last_bot_msg_id": None, "selected_cat_id": None}
        
        pending_order = {"service": row[1], "price": row[2]} if row[1] else None
        return {
            "state": row[0],
            "pending_order": pending_order,
            "last_bot_msg_id": row[3],
            "selected_cat_id": row[4]
        }

def update_user_state(chat_id, state=None, pending_order=None, last_bot_msg_id=None, selected_cat_id=None, clear_pending=False):
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        if state:
            cursor.execute("UPDATE user_states SET state = ? WHERE chat_id = ?", (state, chat_id))
        if last_bot_msg_id is not None:
            cursor.execute("UPDATE user_states SET last_bot_msg_id = ? WHERE chat_id = ?", (last_bot_msg_id, chat_id))
        if selected_cat_id is not None:
            cursor.execute("UPDATE user_states SET selected_cat_id = ? WHERE chat_id = ?", (selected_cat_id, chat_id))
        if pending_order:
            cursor.execute("UPDATE user_states SET pending_service = ?, pending_price = ? WHERE chat_id = ?", 
                           (pending_order["service"], pending_order["price"], chat_id))
        if clear_pending:
            cursor.execute("UPDATE user_states SET pending_service = NULL, pending_price = NULL, selected_cat_id = NULL WHERE chat_id = ?", (chat_id,))
        conn.commit()

def get_user_orders(chat_id):
    with sqlite3.connect(DB_FILE) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT order_id_user AS id, category_id, service, price, time, date FROM orders WHERE chat_id = ? ORDER BY id ASC", (chat_id,))
        rows = cursor.fetchall()
        return [dict(r) for r in rows]

def get_main_menu_markup():
    markup = types.InlineKeyboardMarkup(row_width=2)
    
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, icon FROM categories")
        cats = cursor.fetchall()
        
    buttons = []
    for cat in cats:
        buttons.append(types.InlineKeyboardButton(f"{cat[2]} {cat[1]}", callback_data=f"cat_show_{cat[0]}"))
        
    markup.add(*buttons)
    
    b3 = types.InlineKeyboardButton("📊 التقارير والإحصائيات", callback_data="menu_reports")
    b4 = types.InlineKeyboardButton("💾 النسخ الاحتياطي والضبط", callback_data="backup_menu")
    b5 = types.InlineKeyboardButton("🗑️ حذف آخر طلب", callback_data="confirm_delete")
    
    markup.add(b3, b4)
    markup.add(b5)
    return markup

def get_reports_menu_markup():
    markup = types.InlineKeyboardMarkup(row_width=2)
    b1 = types.InlineKeyboardButton("📆 تقرير اليوم", callback_data="rep_today")
    b2 = types.InlineKeyboardButton("📈 تقرير الشهر", callback_data="rep_month")
    b3 = types.InlineKeyboardButton("📅 الشهور السابقة", callback_data="rep_all_months")
    b4 = types.InlineKeyboardButton("🔙 رجوع للقائمة الرئيسية", callback_data="main_menu")
    
    markup.add(b1, b2)
    markup.add(b3)
    markup.add(b4)
    return markup

def get_main_menu_text(chat_id):
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    orders = get_user_orders(chat_id)
    today_orders = [o for o in orders if o["date"] == today_str]
    
    total_rev = sum(o["price"] for o in today_orders)
    total_count = len(today_orders)
    
    text = f"📊 *نظرة عامة — اليوم*\n"
    text += f"📅 {today_str}\n"
    text += "━━━━━━━━━━━━━━━━━━\n"
    text += f"💰 إجمالي الإيرادات: *{total_rev} ج*\n"
    text += f"📦 عدد الطلبات: *{total_count}*\n\n"
    
    text += "📝 *ملخص الأقسام اليوم:*\n"
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT c.icon, c.name, COUNT(o.id), SUM(o.price)
            FROM orders o
            JOIN categories c ON o.category_id = c.id
            WHERE o.chat_id = ? AND o.date = ?
            GROUP BY c.id
        """, (chat_id, today_str))
        cat_summary = cursor.fetchall()
        
    if not cat_summary:
        text += "  لا توجد مبيعات في الأقسام اليوم بعد.\n\n"
    else:
        for row in cat_summary:
            text += f"  {row[0]} {row[1]} — عدد {row[2]} — *{row[3]} ج*\n"
        text += "\n"
        
    text += "🕐 *آخر 5 تسجيلات:*\n"
    if not orders:
        text += "  مفيش تسجيلات\n"
    else:
        for o in orders[-5:][::-1]:
            text += f"  `#{o['id']}` {o['service']} — *{o['price']}ج* ({o['date']})\n"
            
    text += "\n📌 *اختار الإجراء أو القسم:*"
    return text

def record_order(chat_id, category_id, service, price):
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM orders WHERE chat_id = ?", (chat_id,))
        count = cursor.fetchone()[0]
        next_order_id = count + 1
        
        now_time = datetime.datetime.now().strftime("%H:%M")
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        
        cursor.execute('''
            INSERT INTO orders (chat_id, order_id_user, category_id, service, price, time, date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (chat_id, next_order_id, category_id, service, int(price), now_time, today_str))
        conn.commit()
    
    update_user_state(chat_id, state="main_menu", clear_pending=True)
    return {
        "id": next_order_id,
        "service": service,
        "price": int(price),
        "time": now_time,
        "date": today_str
    }

def auto_return_to_main(chat_id, message_id):
    try:
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=get_main_menu_text(chat_id),
            reply_markup=get_main_menu_markup()
        )
    except Exception:
        pass

def price_timeout_handler(chat_id, message_id, category_id):
    data = get_user_state(chat_id)
    if data["state"] == "awaiting_price" and data["pending_order"] and data["last_bot_msg_id"] == message_id:
        po = data["pending_order"]
        order = record_order(chat_id, category_id, po["service"], po["price"])
        
        text = f"✅ *تم التسجيل تلقائياً #{order['id']}*\n\n"
        text += f"🛠️ {order['service']}\n"
        text += f"💰 {order['price']}ج\n"
        text += f"🕐 {order['time']}\n\n"
        text += "⏳ رجوع تلقائي للقائمة..."
        
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🔙 رجوع", callback_data="main_menu"))
        
        try:
            bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=text, reply_markup=markup)
            scheduler.add_job(auto_return_to_main, 'date', run_date=datetime.datetime.now() + datetime.timedelta(seconds=2), args=[chat_id, message_id])
        except Exception:
            pass

def normalize_date(date_str):
    date_str = date_str.strip().replace('/', '-')
    formats = ('%Y-%m-%d', '%d-%m-%Y', '%d-%m-%y')
    for fmt in formats:
        try:
            return datetime.datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return date_str

def auto_daily_backup():
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    current_month = datetime.date.today().strftime("%Y-%m")
    current_year = datetime.date.today().strftime("%Y")
    
    wb = Workbook()
    
    # --- Styles ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(size=16, bold=True, color="1F4E78")
    money_font = Font(size=11, color="006100")
    money_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_font = Font(color="9C0006")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Fetch all data
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT order_id_user, service, price, time, date, chat_id, category_id FROM orders ORDER BY date DESC, time DESC")
        all_orders = cursor.fetchall()
        cursor.execute("SELECT id, name, icon FROM categories")
        categories = {row[0]: f"{row[2]} {row[1]}" for row in cursor.fetchall()}
        cursor.execute("SELECT id, name, category_id FROM services")
        services_map = {row[0]: row[1] for row in cursor.fetchall()}
    
    today_orders = [o for o in all_orders if o[4] == today_str]
    month_orders = [o for o in all_orders if o[4].startswith(current_month)]
    year_orders = [o for o in all_orders if o[4].startswith(current_year)]
    total_day_rev = sum(o[2] for o in today_orders)
    total_month_rev = sum(o[2] for o in month_orders)
    total_year_rev = sum(o[2] for o in year_orders)
    total_all_rev = sum(o[2] for o in all_orders)
    
    # ============================================
    # SHEET 1: DASHBOARD
    # ============================================
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    
    ws_dash.merge_cells("A1:F1")
    ws_dash["A1"] = f"Wash & Scan - Daily Report ({today_str})"
    ws_dash["A1"].font = title_font
    ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 30
    
    kpi_data = [
        ["Today Revenue", f"{total_day_rev} EGP", len(today_orders), "orders"],
        ["Month Revenue", f"{total_month_rev} EGP", len(month_orders), "orders"],
        ["Year Revenue", f"{total_year_rev} EGP", len(year_orders), "orders"],
        ["All Time Revenue", f"{total_all_rev} EGP", len(all_orders), "orders"]
    ]
    
    ws_dash["A3"] = "Period"
    ws_dash["B3"] = "Revenue"
    ws_dash["C3"] = "Count"
    ws_dash["D3"] = "Unit"
    for col in ["A", "B", "C", "D"]:
        ws_dash[f"{col}3"].fill = header_fill
        ws_dash[f"{col}3"].font = header_font
        ws_dash[f"{col}3"].alignment = center_align
        ws_dash[f"{col}3"].border = thin_border
    
    for idx, row in enumerate(kpi_data, start=4):
        ws_dash[f"A{idx}"] = row[0]
        ws_dash[f"B{idx}"] = row[1]
        ws_dash[f"C{idx}"] = row[2]
        ws_dash[f"D{idx}"] = row[3]
        for col in ["A", "B", "C", "D"]:
            ws_dash[f"{col}{idx}"].border = thin_border
            ws_dash[f"{col}{idx}"].alignment = center_align
        ws_dash[f"B{idx}"].fill = money_fill
        ws_dash[f"B{idx}"].font = money_font
    
    ws_dash["A9"] = "Category Breakdown - Today"
    ws_dash["A9"].font = subheader_font
    ws_dash["A9"].fill = subheader_fill
    ws_dash.merge_cells("A9:D9")
    
    ws_dash["A10"] = "Category"
    ws_dash["B10"] = "Orders"
    ws_dash["C10"] = "Revenue"
    ws_dash["D10"] = "% Share"
    for col in ["A", "B", "C", "D"]:
        ws_dash[f"{col}10"].fill = header_fill
        ws_dash[f"{col}10"].font = header_font
        ws_dash[f"{col}10"].alignment = center_align
        ws_dash[f"{col}10"].border = thin_border
    
    cat_breakdown = {}
    for o in today_orders:
        cat_name = categories.get(o[6], "Unknown")
        if cat_name not in cat_breakdown:
            cat_breakdown[cat_name] = {"count": 0, "revenue": 0}
        cat_breakdown[cat_name]["count"] += 1
        cat_breakdown[cat_name]["revenue"] += o[2]
    
    row_idx = 11
    for cat_name, stats in sorted(cat_breakdown.items(), key=lambda x: x[1]["revenue"], reverse=True):
        pct = (stats["revenue"] / total_day_rev * 100) if total_day_rev > 0 else 0
        ws_dash[f"A{row_idx}"] = cat_name
        ws_dash[f"B{row_idx}"] = stats["count"]
        ws_dash[f"C{row_idx}"] = stats["revenue"]
        ws_dash[f"D{row_idx}"] = f"{pct:.1f}%"
        for col in ["A", "B", "C", "D"]:
            ws_dash[f"{col}{row_idx}"].border = thin_border
            ws_dash[f"{col}{row_idx}"].alignment = center_align
        ws_dash[f"C{row_idx}"].fill = money_fill
        row_idx += 1
    
    ws_dash.column_dimensions["A"].width = 22
    ws_dash.column_dimensions["B"].width = 18
    ws_dash.column_dimensions["C"].width = 12
    ws_dash.column_dimensions["D"].width = 12
    
    # ============================================
    # SHEET 2: ALL ORDERS
    # ============================================
    ws_orders = wb.create_sheet("All Orders")
    
    headers = ["#", "Service", "Category", "Price (EGP)", "Time", "Date", "Order ID"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_orders.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    for row_idx, order in enumerate(all_orders, start=2):
        ws_orders.cell(row=row_idx, column=1, value=row_idx-1)
        ws_orders.cell(row=row_idx, column=2, value=order[1])
        ws_orders.cell(row=row_idx, column=3, value=categories.get(order[6], "Unknown"))
        ws_orders.cell(row=row_idx, column=4, value=order[2])
        ws_orders.cell(row=row_idx, column=5, value=order[3])
        ws_orders.cell(row=row_idx, column=6, value=order[4])
        ws_orders.cell(row=row_idx, column=7, value=order[0])
        for col in range(1, 8):
            ws_orders.cell(row=row_idx, column=col).border = thin_border
            ws_orders.cell(row=row_idx, column=col).alignment = center_align
        ws_orders.cell(row=row_idx, column=4).fill = money_fill
        ws_orders.cell(row=row_idx, column=4).font = money_font
    
    total_row = len(all_orders) + 2
    ws_orders.cell(row=total_row, column=1, value="TOTAL")
    ws_orders.cell(row=total_row, column=4, value=total_all_rev)
    for col in [1, 4]:
        ws_orders.cell(row=total_row, column=col).font = Font(bold=True, size=12)
        ws_orders.cell(row=total_row, column=col).fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        ws_orders.cell(row=total_row, column=col).border = thin_border
    
    ws_orders.column_dimensions["A"].width = 6
    ws_orders.column_dimensions["B"].width = 28
    ws_orders.column_dimensions["C"].width = 22
    ws_orders.column_dimensions["D"].width = 14
    ws_orders.column_dimensions["E"].width = 10
    ws_orders.column_dimensions["F"].width = 12
    ws_orders.column_dimensions["G"].width = 10
    
    # ============================================
    # SHEET 3: TODAY ORDERS
    # ============================================
    ws_today = wb.create_sheet("Today Orders")
    
    ws_today.merge_cells("A1:F1")
    ws_today["A1"] = f"Orders for {today_str}"
    ws_today["A1"].font = title_font
    ws_today["A1"].alignment = center_align
    ws_today.row_dimensions[1].height = 25
    
    headers_today = ["#", "Service", "Category", "Price", "Time", "Order ID"]
    for col_idx, header in enumerate(headers_today, 1):
        cell = ws_today.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    for row_idx, order in enumerate(today_orders, start=4):
        ws_today.cell(row=row_idx, column=1, value=row_idx-3)
        ws_today.cell(row=row_idx, column=2, value=order[1])
        ws_today.cell(row=row_idx, column=3, value=categories.get(order[6], "Unknown"))
        ws_today.cell(row=row_idx, column=4, value=order[2])
        ws_today.cell(row=row_idx, column=5, value=order[3])
        ws_today.cell(row=row_idx, column=6, value=order[0])
        for col in range(1, 7):
            ws_today.cell(row=row_idx, column=col).border = thin_border
            ws_today.cell(row=row_idx, column=col).alignment = center_align
        ws_today.cell(row=row_idx, column=4).fill = money_fill
    
    t_total = len(today_orders) + 4
    ws_today.cell(row=t_total, column=1, value="TOTAL")
    ws_today.cell(row=t_total, column=4, value=total_day_rev)
    for col in [1, 4]:
        ws_today.cell(row=t_total, column=col).font = Font(bold=True, size=12)
        ws_today.cell(row=t_total, column=col).fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        ws_today.cell(row=t_total, column=col).border = thin_border
    
    ws_today.column_dimensions["A"].width = 6
    ws_today.column_dimensions["B"].width = 28
    ws_today.column_dimensions["C"].width = 22
    ws_today.column_dimensions["D"].width = 14
    ws_today.column_dimensions["E"].width = 10
    ws_today.column_dimensions["F"].width = 10
    
    # ============================================
    # SHEET 4: MONTHLY SUMMARY
    # ============================================
    ws_month = wb.create_sheet("Monthly Summary")
    
    ws_month.merge_cells("A1:E1")
    ws_month["A1"] = f"Monthly Summary - {current_year}"
    ws_month["A1"].font = title_font
    ws_month["A1"].alignment = center_align
    
    month_headers = ["Month", "Total Orders", "Revenue (EGP)", "Avg Order", "Trend"]
    for col_idx, header in enumerate(month_headers, 1):
        cell = ws_month.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    monthly_data = {}
    for o in all_orders:
        m = o[4][:7]
        if m not in monthly_data:
            monthly_data[m] = {"count": 0, "revenue": 0}
        monthly_data[m]["count"] += 1
        monthly_data[m]["revenue"] += o[2]
    
    row_idx = 4
    prev_rev = 0
    for month in sorted(monthly_data.keys(), reverse=True):
        stats = monthly_data[month]
        avg = stats["revenue"] // stats["count"] if stats["count"] > 0 else 0
        trend = "UP" if stats["revenue"] > prev_rev else "DOWN" if stats["revenue"] < prev_rev else "FLAT"
        prev_rev = stats["revenue"]
        ws_month.cell(row=row_idx, column=1, value=month)
        ws_month.cell(row=row_idx, column=2, value=stats["count"])
        ws_month.cell(row=row_idx, column=3, value=stats["revenue"])
        ws_month.cell(row=row_idx, column=4, value=avg)
        ws_month.cell(row=row_idx, column=5, value=trend)
        for col in range(1, 6):
            ws_month.cell(row=row_idx, column=col).border = thin_border
            ws_month.cell(row=row_idx, column=col).alignment = center_align
        ws_month.cell(row=row_idx, column=3).fill = money_fill
        if trend == "UP":
            ws_month.cell(row=row_idx, column=5).fill = money_fill
        elif trend == "DOWN":
            ws_month.cell(row=row_idx, column=5).fill = warning_fill
            ws_month.cell(row=row_idx, column=5).font = warning_font
        row_idx += 1
    
    ws_month.column_dimensions["A"].width = 12
    ws_month.column_dimensions["B"].width = 14
    ws_month.column_dimensions["C"].width = 16
    ws_month.column_dimensions["D"].width = 12
    ws_month.column_dimensions["E"].width = 10
    
    # ============================================
    # SHEET 5: TOP SERVICES
    # ============================================
    ws_top = wb.create_sheet("Top Services")
    
    ws_top.merge_cells("A1:D1")
    ws_top["A1"] = "Top Performing Services - All Time"
    ws_top["A1"].font = title_font
    ws_top["A1"].alignment = center_align
    
    top_headers = ["Rank", "Service", "Total Orders", "Total Revenue"]
    for col_idx, header in enumerate(top_headers, 1):
        cell = ws_top.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    service_stats = {}
    for o in all_orders:
        svc = o[1]
        if svc not in service_stats:
            service_stats[svc] = {"count": 0, "revenue": 0}
        service_stats[svc]["count"] += 1
        service_stats[svc]["revenue"] += o[2]
    
    sorted_services = sorted(service_stats.items(), key=lambda x: x[1]["revenue"], reverse=True)
    for rank, (svc, stats) in enumerate(sorted_services[:20], start=1):
        row = rank + 3
        ws_top.cell(row=row, column=1, value=rank)
        ws_top.cell(row=row, column=2, value=svc)
        ws_top.cell(row=row, column=3, value=stats["count"])
        ws_top.cell(row=row, column=4, value=stats["revenue"])
        for col in range(1, 5):
            ws_top.cell(row=row, column=col).border = thin_border
            ws_top.cell(row=row, column=col).alignment = center_align
        ws_top.cell(row=row, column=4).fill = money_fill
        if rank <= 3:
            ws_top.cell(row=row, column=1).fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    
    ws_top.column_dimensions["A"].width = 8
    ws_top.column_dimensions["B"].width = 30
    ws_top.column_dimensions["C"].width = 14
    ws_top.column_dimensions["D"].width = 16
    
    # ============================================
    # SAVE & SEND
    # ============================================
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    bio.name = f"Daily_Report_{today_str}.xlsx"
    
    try:
        report_text = f"📊 *Daily Professional Report*\n"
        report_text += f"📅 Date: `{today_str}`\n"
        report_text += "━━━━━━━━━━━━━━━━━━\n"
        report_text += f"💰 Today: *{total_day_rev} EGP* ({len(today_orders)} orders)\n"
        report_text += f"📈 Month: *{total_month_rev} EGP* ({len(month_orders)} orders)\n"
        report_text += f"📊 Year: *{total_year_rev} EGP* ({len(year_orders)} orders)\n"
        report_text += f"🏆 All Time: *{total_all_rev} EGP* ({len(all_orders)} orders)\n\n"
        report_text += "📎 Attached: 5-sheet Excel workbook with full analytics"
        
        bot.send_document(ADMIN_CHAT_ID, bio, caption=report_text)
        print(f"Daily backup sent: {bio.name}")
    except Exception as e:
        print(f"Failed to send automated backup: {e}")

scheduler.add_job(auto_daily_backup, 'cron', hour=23, minute=59)

# --- Command Handlers ---
@bot.message_handler(commands=['start'])
def start_cmd(message):
    chat_id = message.chat.id
    update_user_state(chat_id, state="main_menu", clear_pending=True, last_bot_msg_id=0)
    bot.send_message(chat_id, get_main_menu_text(chat_id), reply_markup=get_main_menu_markup())

# --- Callback Query Handlers ---
@bot.callback_query_handler(func=lambda call: True)
def handle_callbacks(call):
    chat_id = call.message.chat.id
    msg_id = call.message.message_id
    data = get_user_state(chat_id)
    
    if call.data == "main_menu":
        update_user_state(chat_id, state="main_menu", clear_pending=True)
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=get_main_menu_text(chat_id), reply_markup=get_main_menu_markup())

    elif call.data == "menu_reports":
        update_user_state(chat_id, state="reports_menu")
        text = "📊 *مطبخ التقارير والإحصائيات*\n\nاختار نوع التقرير المطلوب استعراضه:"
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=get_reports_menu_markup())

    elif call.data.startswith("cat_show_"):
        cat_id = int(call.data.split("_")[2])
        update_user_state(chat_id, state="service_list", selected_cat_id=cat_id)
        
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name, icon FROM categories WHERE id = ?", (cat_id,))
            cat_info = cursor.fetchone()
            cursor.execute("SELECT name, price FROM services WHERE category_id = ? AND is_active = 1", (cat_id,))
            services = cursor.fetchall()
            
        title = f"{cat_info[1]} {cat_info[0]}"
        markup = types.InlineKeyboardMarkup(row_width=1)
        
        for srv in services:
            markup.add(types.InlineKeyboardButton(f"{srv[0]} — {srv[1]}ج", callback_data=f"srv_{srv[0]}_{srv[1]}"))
            
        markup.add(types.InlineKeyboardButton("🔙 رجوع للقائمة الرئيسية", callback_data="main_menu"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=f"📍 قسم: *{title}*\n\nاضغط لتسجيل الإجراء مباشرة:", reply_markup=markup)

    elif call.data.startswith("srv_"):
        parts = call.data.split("_")
        service_name = parts[1]
        default_price = int(parts[2])
        cat_id = data["selected_cat_id"]
        
        update_user_state(chat_id, state="awaiting_price", pending_order={"service": service_name, "price": default_price}, last_bot_msg_id=msg_id)
        
        text = f"💰 *تأكيد السعر*\n\n"
        text += f"🛠️ {service_name}\n"
        text += f"السعر الافتراضي: *{default_price}ج*\n\n"
        text += "✅ اضغط تأكيد أو اكتب سعر جديد في الشات خلال 10 ثواني..."
        
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton(f"✅ تأكيد ({default_price}ج)", callback_data="confirm_default"),
            types.InlineKeyboardButton("🔙 رجوع", callback_data="main_menu")
        )
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)
        scheduler.add_job(price_timeout_handler, 'date', run_date=datetime.datetime.now() + datetime.timedelta(seconds=10), args=[chat_id, msg_id, cat_id])

    elif call.data == "confirm_default":
        if data["state"] == "awaiting_price" and data["pending_order"]:
            po = data["pending_order"]
            cat_id = data["selected_cat_id"]
            order = record_order(chat_id, cat_id, po["service"], po["price"])
            
            text = f"✅ *تم التسجيل #{order['id']}*\n\n"
            text += f"🛠️ {order['service']}\n"
            text += f"💰 {order['price']}ج\n"
            text += f"🕐 {order['time']}\n\n"
            text += "⏳ رجوع تلقائي للقائمة..."
            
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("🔙 رجوع", callback_data="main_menu"))
            bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)
            scheduler.add_job(auto_return_to_main, 'date', run_date=datetime.datetime.now() + datetime.timedelta(seconds=2), args=[chat_id, msg_id])

    elif call.data == "backup_menu":
        text = "💾 *إدارة النسخ الاحتياطي والضبط*\n\nتحكم بالبيانات المخزنة أو توجه إلى لوحة التحكم بالأقسام والأسعار:"
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("⚙️ إعدادات الأقسام والأسعار", callback_data="settings_main"),
            types.InlineKeyboardButton("➕ إنشاء قسم جديد تماماً", callback_data="add_category_trigger"),
            types.InlineKeyboardButton("📤 تحميل البيانات فورا (CSV)", callback_data="export_csv"),
            types.InlineKeyboardButton("📥 استعادة البيانات (رفع CSV)", callback_data="import_csv_trigger"),
            types.InlineKeyboardButton("🔙 رجوع للقائمة الرئيسية", callback_data="main_menu")
        )
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data == "settings_main":
        text = "⚙️ *لوحة تحكم الأقسام والخدمات*\n\nاختار القسم لتعديل أسعاره أو إضافة خدمة جديدة داخله:"
        markup = types.InlineKeyboardMarkup(row_width=1)
        
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, name, icon FROM categories")
            cats = cursor.fetchall()
            
        for cat in cats:
            markup.add(types.InlineKeyboardButton(f"إدارة: {cat[2]} {cat[1]}", callback_data=f"setcat_{cat[0]}"))
            
        markup.add(types.InlineKeyboardButton("🔙 رجوع للخلف", callback_data="backup_menu"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data == "add_category_trigger":
        update_user_state(chat_id, state="awaiting_new_category", last_bot_msg_id=msg_id)
        text = "➕ *إنشاء قسم جديد*\n\nاكتب اسم القسم والـ emoji المفضل في رسالة واحدة مثل هذا المثال:\n\n`غسيل سيارات — 🚗`"
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("❌ إلغاء", callback_data="backup_menu"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data.startswith("setcat_"):
        cat_id = int(call.data.split("_")[1])
        update_user_state(chat_id, selected_cat_id=cat_id)
        
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name, icon FROM categories WHERE id = ?", (cat_id,))
            cat_info = cursor.fetchone()
            cursor.execute("SELECT id, name, price FROM services WHERE category_id = ? AND is_active = 1", (cat_id,))
            services = cursor.fetchall()
            
        text = f"⚙️ *تعديل قسم: {cat_info[1]} {cat_info[0]}*\n\n"
        text += "الخدمات الحالية المسجلة:\n"
        if not services:
            text += "  لا توجد خدمات في هذا القسم بعد.\n"
        for s in services:
            text += f"• {s[1]} — {s[2]}ج\n"
            
        text += "\n➕ لإضافة خدمة جديدة لهذا القسم اضغط على الزر بالأسفل."
        
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("➕ إضافة خدمة / إجراء جديد", callback_data="add_service_trigger"),
            types.InlineKeyboardButton("🗑️ حذف القسم", callback_data="remove_this_category"),
            types.InlineKeyboardButton("🗑️ حذف خدمة", callback_data="remove_service_trigger"),
            types.InlineKeyboardButton("🔙 رجوع للضبط", callback_data="settings_main")
        )
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data == "add_service_trigger":
        update_user_state(chat_id, state="awaiting_new_service_name", last_bot_msg_id=msg_id)
        text = "✍️ *إضافة خدمة جديدة*\n\nاكتب اسم الخدمة والسعر في رسالة واحدة بالشات بهذا الشكل تماماً:\n\n`غسيل شامل — 300`\n\n⚠️ تأكد من وضع الشرطة المائلة `—` بين الاسم والسعر."
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("❌ إلغاء", callback_data="settings_main"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data == "remove_this_category":
        cat_id = data["selected_cat_id"]
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name, icon FROM categories WHERE id = ?", (cat_id,))
            cat_info = cursor.fetchone()
        text = f"⚠️ *تأكيد حذف القسم*\n\nهل أنت متأكد من حذف هذا القسم بالكامل؟\n\n📂 {cat_info[1]} {cat_info[0]}\n\n❗ سيتم إلغاء تفعيل جميع الخدمات بداخله. الطلبات السابقة ستبقى محفوظة."
        markup = types.InlineKeyboardMarkup(row_width=2)
        markup.add(
            types.InlineKeyboardButton("✅ نعم، احذف", callback_data="confirm_delcat_now"),
            types.InlineKeyboardButton("❌ لا، تراجع", callback_data=f"setcat_{cat_id}")
        )
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data == "confirm_delcat_now":
        cat_id = data["selected_cat_id"]
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name, icon FROM categories WHERE id = ?", (cat_id,))
            cat_info = cursor.fetchone()
            cursor.execute("UPDATE services SET is_active = 0 WHERE category_id = ?", (cat_id,))
            cursor.execute("DELETE FROM categories WHERE id = ?", (cat_id,))
            conn.commit()
        update_user_state(chat_id, selected_cat_id=None)
        text = f"🗑️ *تم حذف القسم بنجاح!*\n\n📂 {cat_info[1]} {cat_info[0]}\n\n⏳ رجوع تلقائي..."
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🔙 رجوع للضبط", callback_data="settings_main"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)
        scheduler.add_job(auto_return_to_main, 'date', run_date=datetime.datetime.now() + datetime.timedelta(seconds=2), args=[chat_id, msg_id])

    elif call.data == "remove_service_trigger":
        cat_id = data["selected_cat_id"]
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name, icon FROM categories WHERE id = ?", (cat_id,))
            cat_info = cursor.fetchone()
            cursor.execute("SELECT id, name, price FROM services WHERE category_id = ? AND is_active = 1", (cat_id,))
            services = cursor.fetchall()
            
        text = f"🗑️ *حذف خدمة من قسم: {cat_info[1]} {cat_info[0]}*\n\nاختار الخدمة المراد حذفها:"
        markup = types.InlineKeyboardMarkup(row_width=1)
        
        if not services:
            text = "❌ لا توجد خدمات في هذا القسم لحذفها."
            markup.add(types.InlineKeyboardButton("🔙 رجوع", callback_data="settings_main"))
            bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)
        else:
            for srv in services:
                markup.add(types.InlineKeyboardButton(f"🗑️ {srv[1]} — {srv[2]}ج", callback_data=f"delsrv_{srv[0]}"))
            markup.add(types.InlineKeyboardButton("❌ إلغاء", callback_data="settings_main"))
            bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data.startswith("delsrv_"):
        srv_id = int(call.data.split("_")[1])
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name, price FROM services WHERE id = ?", (srv_id,))
            srv_info = cursor.fetchone()
            cursor.execute("UPDATE services SET is_active = 0 WHERE id = ?", (srv_id,))
            conn.commit()
        text = f"🗑️ *تم حذف الخدمة بنجاح!*\n\n🛠️ {srv_info[0]} — {srv_info[1]}ج\n\n⏳ رجوع تلقائي..."
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🔙 رجوع", callback_data="settings_main"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)
        scheduler.add_job(auto_return_to_main, 'date', run_date=datetime.datetime.now() + datetime.timedelta(seconds=2), args=[chat_id, msg_id])

    elif call.data == "rep_today":
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        orders = get_user_orders(chat_id)
        today_orders = [o for o in orders if o["date"] == today_str]
        total_rev = sum(o["price"] for o in today_orders)
        
        text = f"📆 *تقرير اليوم بالتفصيل*\n"
        text += f"📅 {today_str}\n"
        text += "━━━━━━━━━━━━━━━━━━\n"
        text += f"💰 الإجمالي: *{total_rev} ج*\n"
        text += f"📦 الطلبات: *{len(today_orders)}*\n\n"
        text += "📝 *التفاصيل:*\n"
        
        if not today_orders:
            text += "  مفيش طلبات مسجلة اليوم حتى الآن."
        else:
            for o in today_orders:
                text += f"  `#{o['id']}` {o['service']} — *{o['price']}ج* ({o['time']})\n"
                
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🔙 رجوع للتقارير", callback_data="menu_reports"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data == "rep_month":
        current_month = datetime.date.today().strftime("%Y-%m")
        orders = get_user_orders(chat_id)
        month_orders = [o for o in orders if o["date"].startswith(current_month)]
        total_rev = sum(o["price"] for o in month_orders)
        
        text = f"📈 *تقرير إحصائيات الشهر الحالي*\n"
        text += f"📅 {current_month}\n"
        text += "━━━━━━━━━━━━━━━━━━\n"
        text += f"💰 إجمالي الإيرادات: *{total_rev} ج*\n"
        text += f"📦 إجمالي الطلبات: *{len(month_orders)}*\n\n"
        
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🔙 رجوع للتقارير", callback_data="menu_reports"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data == "rep_all_months":
        text = f"📅 *الملخص السنوي / الشهور السابقة*\n"
        text += "━━━━━━━━━━━━━━━━━━\n"
        orders = get_user_orders(chat_id)
        monthly_summaries = {}
        for o in orders:
            month_key = o["date"][:7]
            monthly_summaries[month_key] = monthly_summaries.get(month_key, 0) + o["price"]
            
        if not monthly_summaries:
            text += "لا يوجد مبيعات مؤرشفة لشهور سابقة بعد."
        else:
            for m_key, m_total in sorted(monthly_summaries.items(), reverse=True):
                text += f"📅 شهر *{m_key}* — إجمالي الإيراد: *{m_total} ج*\n"
                
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🔙 رجوع للتقارير", callback_data="menu_reports"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data == "confirm_delete":
        orders = get_user_orders(chat_id)
        if not orders:
            bot.answer_callback_query(call.id, "مفيش طلبات عشان تحذفها!", show_alert=True)
            return
        
        last_order = orders[-1]
        text = f"🗑️ *تأكيد الحذف*\n\nهل أنت متأكد من حذف آخر طلب؟\n\n`#{last_order['id']}` {last_order['service']} — *{last_order['price']}ج*"
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("⚠️ نعم، احذف", callback_data="delete_execute"),
            types.InlineKeyboardButton("❌ لا، تراجع", callback_data="main_menu")
        )
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)

    elif call.data == "delete_execute":
        orders = get_user_orders(chat_id)
        if orders:
            last_order = orders[-1]
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM orders WHERE chat_id = ? AND order_id_user = ?", (chat_id, last_order["id"]))
                conn.commit()
            
            text = "🗑️ تم الحذف بنجاح... رجوع تلقائي"
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("🔙 رجوع", callback_data="main_menu"))
            bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)
            scheduler.add_job(auto_return_to_main, 'date', run_date=datetime.datetime.now() + datetime.timedelta(seconds=2), args=[chat_id, msg_id])

    elif call.data == "export_csv":
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        orders = get_user_orders(chat_id)
        csv_buffer = io.StringIO()
        csv_buffer.write('\ufeff')
        writer = csv.writer(csv_buffer)
        writer.writerow(["رقم الطلب", "نوع الخدمة", "السعر", "الوقت", "التاريخ"])
        for o in orders:
            writer.writerow([o["id"], o["service"], o["price"], o["time"], o["date"]])
        csv_buffer.seek(0)
        bio = io.BytesIO(csv_buffer.getvalue().encode('utf-8'))
        bio.name = f"تقرير_الشغل_{chat_id}_{today_str}.csv"
        try:
            bot.send_document(chat_id, bio, caption="📤 تم سحب النسخة بنجاح.")
        except Exception:
            pass

    elif call.data == "import_csv_trigger":
        update_user_state(chat_id, state="awaiting_restore_file", last_bot_msg_id=msg_id)
        text = "📥 قم بإرسال ملف الـ CSV لاستعادة البيانات الحالية واستبدالها..."
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("❌ إلغاء", callback_data="backup_menu"))
        bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=text, reply_markup=markup)


# --- Message Handling ---
@bot.message_handler(content_types=['text', 'document'])
def handle_all_messages(message):
    chat_id = message.chat.id
    data = get_user_state(chat_id)
    target_msg_id = data["last_bot_msg_id"]
    
    if data["state"] == "awaiting_new_category" and message.content_type == 'text':
        text_received = message.text.strip()
        try: bot.delete_message(chat_id, message.message_id)
        except Exception: pass
        
        if "—" in text_received:
            try:
                cat_name, cat_icon = text_received.split("—")
                cat_name = cat_name.strip()
                cat_icon = cat_icon.strip()
                
                with sqlite3.connect(DB_FILE) as conn:
                    cursor = conn.cursor()
                    cursor.execute("INSERT INTO categories (name, icon) VALUES (?, ?)", (cat_name, cat_icon))
                    conn.commit()
                text = f"✅ *تم إنشاء القسم الجديد بنجاح!*\n\n📂 القسم: {cat_icon} {cat_name}\n📌 يمكنك الآن التوجه للإعدادات وإضافة خدمات له."
            except Exception as e:
                text = f"❌ حدث خطأ، يرجى التأكد من عدم تكرار اسم القسم.\n`{str(e)}`"
        else:
            text = "❌ صيغة غير صحيحة. اكتبها كالتالي: `غسيل سيارات — 🚗`"
            
        update_user_state(chat_id, state="main_menu")
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("⚙️ إعدادات الأقسام والأسعار", callback_data="settings_main"))
        bot.send_message(chat_id, text, reply_markup=markup)
        return

    elif data["state"] == "awaiting_new_service_name" and message.content_type == 'text':
        text_received = message.text.strip()
        try: bot.delete_message(chat_id, message.message_id)
        except Exception: pass
            
        if "—" in text_received:
            try:
                srv_name, srv_price = text_received.split("—")
                srv_name = srv_name.strip()
                srv_price = int(srv_price.strip())
                cat_id = data["selected_cat_id"]
                
                if cat_id:
                    with sqlite3.connect(DB_FILE) as conn:
                        cursor = conn.cursor()
                        cursor.execute("INSERT INTO services (category_id, name, price) VALUES (?, ?, ?)", (cat_id, srv_name, srv_price))
                        conn.commit()
                    text = f"✅ *تمت إضافة الإجراء الجديد بنجاح!*\n\n🛠️ الخدمة: {srv_name}\n💰 السعر: {srv_price}ج"
                else:
                    text = "❌ خطأ في تحديد القسم الحالي المسؤول."
            except Exception as e:
                text = f"❌ خطأ في معالجة المدخلات، تأكد من الصيغة الرقمية الصحيحة.\n`{str(e)}`"
        else:
            text = "❌ صيغة غير صحيحة. يجب كتابتها هكذا: `غسيل شامل — 300`."
            
        update_user_state(chat_id, state="main_menu")
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("⚙️ العودة للضبط", callback_data="settings_main"))
        bot.send_message(chat_id, text, reply_markup=markup)
        return

    elif data["state"] == "awaiting_restore_file":
        if message.content_type == 'document' and message.document.file_name.endswith('.csv'):
            try:
                file_info = bot.get_file(message.document.file_id)
                downloaded_file = bot.download_file(file_info.file_path)
                csv_content = downloaded_file.decode('utf-8-sig', errors='ignore')
                csv_file = io.StringIO(csv_content)
                reader = csv.reader(csv_file)
                header = next(reader, None)
                
                if not header or "نوع الخدمة" not in header or "السعر" not in header:
                    raise ValueError("الملف المرفوع لا يطابق الهيكلية المطلوبة.")
                
                valid_rows = []
                for row in reader:
                    if len(row) >= 5:
                        clean_date = normalize_date(row[4])
                        valid_rows.append((chat_id, int(row[0]), 1, row[1], int(row[2]), row[3], clean_date))
                
                if valid_rows:
                    with sqlite3.connect(DB_FILE) as conn:
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM orders WHERE chat_id = ?", (chat_id,))
                        cursor.executemany('''
                            INSERT INTO orders (chat_id, order_id_user, category_id, service, price, time, date)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', valid_rows)
                        conn.commit()
                    text = f"✨ *تم استعادة البيانات لعدد {len(valid_rows)} طلب!*"
                else:
                    text = "❌ الملف المرفوع لا يحتوي على سجلات صالحة."
            except Exception as e:
                text = f"❌ *فشلت عملية الاستعادة:*\n`{str(e)}`"
            
            try: bot.delete_message(chat_id, message.message_id)
            except Exception: pass
                
            update_user_state(chat_id, state="main_menu")
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("🔙 القائمة الرئيسية", callback_data="main_menu"))
            bot.send_message(chat_id, text, reply_markup=markup)
            return

    elif data["state"] == "awaiting_price" and data["pending_order"] and message.content_type == 'text':
        text_clean = message.text.strip()
        if text_clean.isdigit():
            custom_price = int(text_clean)
            po = data["pending_order"]
            cat_id = data["selected_cat_id"]
            order = record_order(chat_id, cat_id, po["service"], custom_price)
            
            try: bot.delete_message(chat_id, message.message_id)
            except Exception: pass
                
            text = f"✅ *تم التسجيل #{order['id']}*\n\n🛠️ {order['service']}\n💰 {order['price']}ج\n\n⏳ رجوع تلقائي..."
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("🔙 رجوع", callback_data="main_menu"))
            
            try:
                bot.edit_message_text(chat_id=chat_id, message_id=target_msg_id, text=text, reply_markup=markup)
                scheduler.add_job(auto_return_to_main, 'date', run_date=datetime.datetime.now() + datetime.timedelta(seconds=2), args=[chat_id, target_msg_id])
            except Exception:
                sent = bot.send_message(chat_id, text, reply_markup=markup)
                scheduler.add_job(auto_return_to_main, 'date', run_date=datetime.datetime.now() + datetime.timedelta(seconds=2), args=[chat_id, sent.message_id])
            return

    try: bot.delete_message(chat_id, message.message_id)
    except Exception: pass

# --- Start Bot ---
if __name__ == '__main__':
    print("Bot is running...")
    bot.infinity_polling()
